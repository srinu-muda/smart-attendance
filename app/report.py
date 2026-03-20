# app/report.py
# Module 6: Report Generator
# Exports attendance as Excel / CSV with full student details
# Uses pandas + openpyxl — no extra installs needed

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from datetime import datetime
from app.models import Session, AttendanceLog, Student
from app.attendance import get_session_report

# ── Output folder for reports ─────────────────────────────────────
REPORTS_DIR = "data/reports"


def ensure_reports_dir():
    """Create reports folder if it doesn't exist."""
    os.makedirs(REPORTS_DIR, exist_ok=True)


def generate_excel(session_id):
    """
    Generate a formatted Excel report for a session.

    Creates two sheets:
      Sheet 1 — Attendance (present + absent with details)
      Sheet 2 — Summary (stats, subject, faculty, time)

    Args:
        session_id : integer session ID

    Returns:
        filepath of generated Excel file
        or None on error
    """
    ensure_reports_dir()

    # Get full report data
    report = get_session_report(session_id)
    if "error" in report:
        print(f"[ERROR] {report['error']}")
        return None

    session   = report["session"]
    present   = report["present"]
    absent    = report["absent"]
    all_rows  = present + absent

    # ── Build main attendance DataFrame ──────────────────────────
    rows = []
    for i, s in enumerate(all_rows, start=1):
        rows.append({
            "S.No"       : i,
            "Student ID" : s["student_id"],
            "Name"       : s["name"],
            "Status"     : s["status"],
            "Time In"    : s["time"],
            "Confidence" : f"{s['confidence']:.0%}"
                           if s["confidence"] > 0 else "—",
        })

    df_attendance = pd.DataFrame(rows)

    # ── Build summary DataFrame ───────────────────────────────────
    summary_rows = [
        ["Subject Code",   session["subject_code"]],
        ["Subject Name",   session["subject_name"]],
        ["Faculty",        session["faculty_name"]],
        ["Date",           session["start_time"].split(" ")[0]],
        ["Start Time",     session["start_time"].split(" ")[1]],
        ["End Time",       session["end_time"].split(" ")[1]
                           if session["end_time"] else "Active"],
        ["Total Students", report["total"]],
        ["Present",        report["total_present"]],
        ["Absent",         report["total_absent"]],
        ["Attendance %",   f"{report['total_present'] / report['total'] * 100:.1f}%"
                           if report["total"] > 0 else "0%"],
        ["Generated At",   datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    df_summary = pd.DataFrame(summary_rows,
                               columns=["Field", "Value"])

    # ── Write to Excel with formatting ───────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"attendance_{session['subject_code']}_" \
                f"session{session_id}_{timestamp}.xlsx"
    filepath  = os.path.join(REPORTS_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        # Sheet 1 — Attendance
        df_attendance.to_excel(writer,
                                sheet_name="Attendance",
                                index=False)

        # Sheet 2 — Summary
        df_summary.to_excel(writer,
                             sheet_name="Summary",
                             index=False)

        # ── Apply formatting ──────────────────────────────────────
        wb = writer.book

        # Format Sheet 1
        ws1 = writer.sheets["Attendance"]
        _format_attendance_sheet(ws1, df_attendance, present)

        # Format Sheet 2
        ws2 = writer.sheets["Summary"]
        _format_summary_sheet(ws2)

    print(f"[✓] Excel report saved → {filepath}")
    return filepath


def _format_attendance_sheet(ws, df, present_list):
    """Apply colors and column widths to attendance sheet."""
    from openpyxl.styles import (PatternFill, Font,
                                  Alignment, Border, Side)

    present_ids = {s["student_id"] for s in present_list}

    # Colors
    header_fill  = PatternFill("solid", fgColor="1F3A8A")
    present_fill = PatternFill("solid", fgColor="C6EFCE")
    absent_fill  = PatternFill("solid", fgColor="FFC7CE")
    alt_fill     = PatternFill("solid", fgColor="EBF2FA")

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    present_font = Font(color="276221", bold=True)
    absent_font  = Font(color="9C0006", bold=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin,
                    top=thin, bottom=thin)

    # Header row
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center",
                                    vertical="center")
        cell.border    = border

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(
            min_row=2, max_row=ws.max_row), start=2):

        # Get student ID from column B (index 1)
        sid    = row[1].value
        status = row[3].value   # Status column

        for cell in row:
            cell.border    = border
            cell.alignment = Alignment(horizontal="center",
                                        vertical="center")

            if status == "PRESENT":
                cell.fill = present_fill
                if cell.column == 4:   # Status column
                    cell.font = present_font
            elif status == "ABSENT":
                cell.fill = absent_fill
                if cell.column == 4:
                    cell.font = absent_font
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    col_widths = [8, 15, 28, 12, 12, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[
            ws.cell(1, i).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


def _format_summary_sheet(ws):
    """Apply formatting to summary sheet."""
    from openpyxl.styles import PatternFill, Font, Alignment

    header_fill = PatternFill("solid", fgColor="1F3A8A")
    label_fill  = PatternFill("solid", fgColor="EBF2FA")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    label_font  = Font(bold=True, color="1F3A8A")

    # Header
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].fill = label_fill
        row[0].font = label_font
        for cell in row:
            cell.alignment = Alignment(horizontal="left",
                                        vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30


def generate_csv(session_id):
    """
    Generate a simple CSV report for a session.

    Args:
        session_id : integer session ID

    Returns:
        filepath of generated CSV file
    """
    ensure_reports_dir()

    report = get_session_report(session_id)
    if "error" in report:
        print(f"[ERROR] {report['error']}")
        return None

    session  = report["session"]
    all_rows = report["present"] + report["absent"]

    rows = []
    for i, s in enumerate(all_rows, start=1):
        rows.append({
            "S.No"       : i,
            "Student ID" : s["student_id"],
            "Name"       : s["name"],
            "Status"     : s["status"],
            "Time In"    : s["time"],
            "Confidence" : round(s["confidence"], 3),
            "Subject"    : session["subject_code"],
            "Date"       : session["start_time"].split(" ")[0],
        })

    df = pd.DataFrame(rows)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"attendance_{session['subject_code']}_" \
                f"session{session_id}_{timestamp}.csv"
    filepath  = os.path.join(REPORTS_DIR, filename)

    df.to_csv(filepath, index=False)
    print(f"[✓] CSV report saved → {filepath}")
    return filepath


def list_reports():
    """List all generated report files."""
    ensure_reports_dir()
    files = os.listdir(REPORTS_DIR)
    if not files:
        print("[INFO] No reports generated yet.")
        return []

    print(f"\n{'─'*55}")
    print(f"  Generated Reports ({len(files)} total)")
    print(f"{'─'*55}")
    for f in sorted(files, reverse=True):
        path = os.path.join(REPORTS_DIR, f)
        size = os.path.getsize(path)
        print(f"  {f}  ({size/1024:.1f} KB)")
    print(f"{'─'*55}")
    return files